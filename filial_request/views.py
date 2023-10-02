from django.shortcuts import render
from .forms import RequestForm
from django.http import HttpResponse
from userprofile.models import User
from openpyxl import load_workbook

# Create your views here.
def request_page(request):
    if request.method=='POST':
        form = RequestForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
        return HttpResponse("Запрос отправлен")
    else:
        form = RequestForm()
        return render(request,'filial/request.html', {'form':form})


def add_students(request):
    if request.method == 'POST':
        workbook = load_workbook(request.FILES['document'])
        worksheet = workbook.active

        student_region = request.user.region
        school_region = request.user.school

        for row in worksheet.iter_rows():
            username = row[0].value
            password = row[0].value
            first_name = row[1].value
            last_name = row[2].value
            class_name = row[3].value

            if User.objects.filter(username=username).exists():
                continue

            user_class_name = username
            user_class_name = ''.join(filter(str.isdigit, user_class_name))

            user_class_password = password
            user_class_password = ''.join(filter(str.isdigit, user_class_password))

            user = User.objects.create_user(
                username=user_class_name,
                password=user_class_password,
                first_name=first_name,
                last_name=last_name,
                class_name = class_name,
                region=student_region,
                school=school_region
            )
            user.save()

        return HttpResponse("Ученики добавлены")
    else:
        return render(request,'filial/addstudent.html')